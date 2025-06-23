import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

#Generating a list of all upercase letters
upper = []
currASCII = 65

for i in range(25):
    upper.append(chr(currASCII))
    currASCII += 1

#Create an empty Excel file with one sheet (xlsxwriter)
writer = pd.ExcelWriter('portfolio.xlsx', engine='xlsxwriter')
emptyDF = pd.DataFrame()
emptyDF.to_excel(writer, sheet_name='TRANSACTIONS-2025', index=False)
writer.close()

#Define your desired columns
columns = [
    "TYPE", "NAME", "ACCOUNT TYPE", "PLATFORM", "DATE", "TIME (UTC)", "TIME (UK)", "ACTION",
    "CURRENCY", "AMOUNT PER UNIT", "QUANTITY",
    "FEES", "COMPANY SECTOR", "TOTAL PRICE/COST (WITHOUT FEES)",
    "TOTAL CONSIDERATION", "PROFIT/LOSS (APPROX)", "NOTES"
]

#Load the existing file (using pandas + openpyxl engine)
file = pd.read_excel('portfolio.xlsx', sheet_name='TRANSACTIONS-2025', engine='openpyxl')

#Add any missing columns (with empty values)
for col in columns:
    if col not in file.columns:
        file[col] = None

#Reorder columns to match your desired order
file = file[columns]

#Save the DataFrame back to Excel (overwrite the sheet)
file.to_excel('portfolio.xlsx', sheet_name='TRANSACTIONS-2025', index=False, engine='openpyxl')

#Load workbook using openpyxl for styling
wb = load_workbook('portfolio.xlsx')
ws = wb['TRANSACTIONS-2025']

#Wrap text in all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

#Set width of each column
for letter in range(len(upper)-1):
    ws.column_dimensions[upper[letter]].width = 30

#Save the styled workbook
wb.save('portfolio.xlsx')